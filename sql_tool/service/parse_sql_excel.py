from openpyxl.reader.excel import load_workbook
import re
from sql_tool.dto.Table import *
from sql_tool.dto.TableColumn import *


def read_excel(file_name):
    # open file default can read and write
    wb = load_workbook(file_name)

    # only read open
    # wb = load_workbook(file_name,read_only=True)

    #  get all sheets
    sheet_name_list = wb.get_sheet_names()
    tables = []
    table = []
    for sheet_name in sheet_name_list:
        sheet_name_temp = wb.get_sheet_by_name(sheet_name)
        max_row = sheet_name_temp.max_row
        index = 1
        rows = sheet_name_temp.rows
        for row in rows:
            index = index + 1
            line = [cell.value for cell in row]
            if is_desc_row(line):
                continue
            if index == max_row:
                tables.append(table)
            if is_end_row(line):
                table_temp = table
                if table_temp:
                    tables.append(table_temp)
                table = []
                continue
            table.append(line)
    return tables


def is_first_row(line):
    if line[0] and line[1] and re.match(r'[\u4e00-\u9fa5]', line[1]) and line[2] is None:
        return True


def is_desc_row(line):
    if line[0] and line[1] and re.match(r'[\u4e00-\u9fa5]', line[0]) and re.match(r'[\u4e00-\u9fa5]', line[1]):
        return True


def is_end_row(line):
    if line[0] is None and line[1] is None:
        return True


def build_table_obj(tables):
    table_objs = []
    for table in tables:
        table_obj = None
        table_clos = []
        for line in table:
            if is_first_row(line):
                table_obj = Table(line[0], line[1])
            else:
                table_clo = TableColumn(line[0], line[1], line[2], line[3], line[4], line[5], line[6], line[7])
                table_clos.append(table_clo);
        table_obj.fieldColumns.extend(table_clos)
        table_objs.append(table_obj)
    return table_objs
